#!/usr/bin/env python3
"""
Iowa Absentee Ballot Breakdown
Inputs  (auto-detected in script folder):
  Absentee County *.pdf          — Iowa SOS county absentee data
  Absentee Congressional *.pdf   — Iowa SOS official CD totals
  voters by county*.pdf          — Iowa SOS voter registration by county
Outputs (same folder, named after the county PDF):
  {stem}_Breakdown.xlsx          — 4-sheet Excel workbook
  {stem}_map.html                — interactive Iowa county choropleth

Run directly:  python3 breakdown.py
Auto-watch:    python3 watcher.py
"""

import json
import ssl
from pathlib import Path
from urllib.request import urlopen

import pandas as pd
import plotly.graph_objects as go
import pdfplumber
from openpyxl.styles import Alignment, Font, PatternFill

# ─────────────────────────────────────────────────────────────────────────────
# MAPPINGS
# ─────────────────────────────────────────────────────────────────────────────

COUNTY_TO_CD = {
    "Allamakee": "CD1", "Benton": "CD1", "Black Hawk": "CD1", "Bremer": "CD1",
    "Buchanan": "CD1", "Butler": "CD1", "Chickasaw": "CD1", "Clayton": "CD1",
    "Delaware": "CD1", "Dubuque": "CD1", "Fayette": "CD1", "Floyd": "CD1",
    "Howard": "CD1", "Iowa": "CD1", "Jackson": "CD1", "Jones": "CD1",
    "Linn": "CD1", "Marshall": "CD1", "Mitchell": "CD1", "Tama": "CD1",
    "Winneshiek": "CD1", "Worth": "CD1",
    "Appanoose": "CD2", "Cedar": "CD2", "Clinton": "CD2", "Davis": "CD2",
    "Des Moines": "CD2", "Henry": "CD2", "Jefferson": "CD2", "Johnson": "CD2",
    "Keokuk": "CD2", "Lee": "CD2", "Louisa": "CD2", "Mahaska": "CD2",
    "Muscatine": "CD2", "Poweshiek": "CD2", "Scott": "CD2", "Van Buren": "CD2",
    "Wapello": "CD2", "Washington": "CD2",
    "Adair": "CD3", "Adams": "CD3", "Boone": "CD3", "Cass": "CD3",
    "Clarke": "CD3", "Dallas": "CD3", "Decatur": "CD3", "Fremont": "CD3",
    "Greene": "CD3", "Guthrie": "CD3", "Harrison": "CD3", "Jasper": "CD3",
    "Lucas": "CD3", "Madison": "CD3", "Marion": "CD3", "Mills": "CD3",
    "Monroe": "CD3", "Montgomery": "CD3", "Page": "CD3", "Polk": "CD3",
    "Pottawattamie": "CD3", "Ringgold": "CD3", "Story": "CD3", "Taylor": "CD3",
    "Union": "CD3", "Warren": "CD3", "Wayne": "CD3",
    "Audubon": "CD4", "Buena Vista": "CD4", "Calhoun": "CD4", "Carroll": "CD4",
    "Cerro Gordo": "CD4", "Cherokee": "CD4", "Clay": "CD4", "Crawford": "CD4",
    "Dickinson": "CD4", "Emmet": "CD4", "Franklin": "CD4", "Grundy": "CD4",
    "Hamilton": "CD4", "Hancock": "CD4", "Hardin": "CD4", "Humboldt": "CD4",
    "Ida": "CD4", "Kossuth": "CD4", "Lyon": "CD4", "Monona": "CD4",
    "O'Brien": "CD4", "Osceola": "CD4", "Palo Alto": "CD4", "Plymouth": "CD4",
    "Pocahontas": "CD4", "Sac": "CD4", "Shelby": "CD4", "Sioux": "CD4",
    "Webster": "CD4", "Winnebago": "CD4", "Woodbury": "CD4", "Wright": "CD4",
}

COUNTY_TO_MARKET = {
    # Cedar Rapids (21)
    "Allamakee": "Cedar Rapids", "Benton": "Cedar Rapids",    "Black Hawk": "Cedar Rapids",
    "Bremer": "Cedar Rapids",    "Buchanan": "Cedar Rapids",  "Butler": "Cedar Rapids",
    "Cedar": "Cedar Rapids",     "Chickasaw": "Cedar Rapids", "Clayton": "Cedar Rapids",
    "Delaware": "Cedar Rapids",  "Dubuque": "Cedar Rapids",   "Fayette": "Cedar Rapids",
    "Grundy": "Cedar Rapids",    "Iowa": "Cedar Rapids",      "Johnson": "Cedar Rapids",
    "Jones": "Cedar Rapids",     "Keokuk": "Cedar Rapids",    "Linn": "Cedar Rapids",
    "Tama": "Cedar Rapids",      "Washington": "Cedar Rapids","Winneshiek": "Cedar Rapids",
    # Davenport (7)
    "Clinton": "Davenport",   "Des Moines": "Davenport", "Henry": "Davenport",
    "Jackson": "Davenport",   "Louisa": "Davenport",     "Muscatine": "Davenport",
    "Scott": "Davenport",
    # Des Moines (35)
    "Adair": "Des Moines",      "Adams": "Des Moines",     "Appanoose": "Des Moines",
    "Audubon": "Des Moines",    "Boone": "Des Moines",     "Calhoun": "Des Moines",
    "Carroll": "Des Moines",    "Clarke": "Des Moines",    "Dallas": "Des Moines",
    "Decatur": "Des Moines",    "Franklin": "Des Moines",  "Greene": "Des Moines",
    "Guthrie": "Des Moines",    "Hamilton": "Des Moines",  "Hardin": "Des Moines",
    "Humboldt": "Des Moines",   "Jasper": "Des Moines",    "Kossuth": "Des Moines",
    "Lucas": "Des Moines",      "Madison": "Des Moines",   "Mahaska": "Des Moines",
    "Marion": "Des Moines",     "Marshall": "Des Moines",  "Monroe": "Des Moines",
    "Pocahontas": "Des Moines", "Polk": "Des Moines",      "Poweshiek": "Des Moines",
    "Ringgold": "Des Moines",   "Story": "Des Moines",     "Taylor": "Des Moines",
    "Union": "Des Moines",      "Warren": "Des Moines",    "Wayne": "Des Moines",
    "Webster": "Des Moines",    "Wright": "Des Moines",
    # Mason City (7)
    "Cerro Gordo": "Mason City", "Floyd": "Mason City",    "Hancock": "Mason City",
    "Howard": "Mason City",      "Mitchell": "Mason City", "Winnebago": "Mason City",
    "Worth": "Mason City",
    # Omaha (9)
    "Cass": "Omaha",         "Crawford": "Omaha",      "Fremont": "Omaha",
    "Harrison": "Omaha",     "Mills": "Omaha",          "Montgomery": "Omaha",
    "Page": "Omaha",         "Pottawattamie": "Omaha",  "Shelby": "Omaha",
    # Ottumwa (4)
    "Davis": "Ottumwa",  "Jefferson": "Ottumwa", "Van Buren": "Ottumwa",
    "Wapello": "Ottumwa",
    # Quincy (1)
    "Lee": "Quincy",
    # Sioux City (13)
    "Buena Vista": "Sioux City", "Cherokee": "Sioux City", "Clay": "Sioux City",
    "Dickinson": "Sioux City",   "Emmet": "Sioux City",    "Ida": "Sioux City",
    "Monona": "Sioux City",      "O'Brien": "Sioux City",  "Palo Alto": "Sioux City",
    "Plymouth": "Sioux City",    "Sac": "Sioux City",      "Sioux": "Sioux City",
    "Woodbury": "Sioux City",
    # Sioux Falls (2)
    "Lyon": "Sioux Falls", "Osceola": "Sioux Falls",
}

# Iowa county FIPS codes (19001–19197, alphabetical)
COUNTY_TO_FIPS = {
    "Adair": "19001", "Adams": "19003", "Allamakee": "19005", "Appanoose": "19007",
    "Audubon": "19009", "Benton": "19011", "Black Hawk": "19013", "Boone": "19015",
    "Bremer": "19017", "Buchanan": "19019", "Buena Vista": "19021", "Butler": "19023",
    "Calhoun": "19025", "Carroll": "19027", "Cass": "19029", "Cedar": "19031",
    "Cerro Gordo": "19033", "Cherokee": "19035", "Chickasaw": "19037", "Clarke": "19039",
    "Clay": "19041", "Clayton": "19043", "Clinton": "19045", "Crawford": "19047",
    "Dallas": "19049", "Davis": "19051", "Decatur": "19053", "Delaware": "19055",
    "Des Moines": "19057", "Dickinson": "19059", "Dubuque": "19061", "Emmet": "19063",
    "Fayette": "19065", "Floyd": "19067", "Franklin": "19069", "Fremont": "19071",
    "Greene": "19073", "Grundy": "19075", "Guthrie": "19077", "Hamilton": "19079",
    "Hancock": "19081", "Hardin": "19083", "Harrison": "19085", "Henry": "19087",
    "Howard": "19089", "Humboldt": "19091", "Ida": "19093", "Iowa": "19095",
    "Jackson": "19097", "Jasper": "19099", "Jefferson": "19101", "Johnson": "19103",
    "Jones": "19105", "Keokuk": "19107", "Kossuth": "19109", "Lee": "19111",
    "Linn": "19113", "Louisa": "19115", "Lucas": "19117", "Lyon": "19119",
    "Madison": "19121", "Mahaska": "19123", "Marion": "19125", "Marshall": "19127",
    "Mills": "19129", "Mitchell": "19131", "Monona": "19133", "Monroe": "19135",
    "Montgomery": "19137", "Muscatine": "19139", "O'Brien": "19141", "Osceola": "19143",
    "Page": "19145", "Palo Alto": "19147", "Plymouth": "19149", "Pocahontas": "19151",
    "Polk": "19153", "Pottawattamie": "19155", "Poweshiek": "19157", "Ringgold": "19159",
    "Sac": "19161", "Scott": "19163", "Shelby": "19165", "Sioux": "19167",
    "Story": "19169", "Tama": "19171", "Taylor": "19173", "Union": "19175",
    "Van Buren": "19177", "Wapello": "19179", "Warren": "19181", "Washington": "19183",
    "Wayne": "19185", "Webster": "19187", "Winnebago": "19189", "Winneshiek": "19191",
    "Woodbury": "19193", "Worth": "19195", "Wright": "19197",
}

IOWA_COUNTIES = set(COUNTY_TO_CD.keys())
_MAIN_PARTIES = {"Democrat", "Republican"}
_ALL_PARTIES  = {"Democrat", "Republican", "No Party", "Other"}

# ─────────────────────────────────────────────────────────────────────────────
# FILE DETECTION
# ─────────────────────────────────────────────────────────────────────────────

def _latest(folder: Path, glob: str) -> Path | None:
    matches = sorted(folder.glob(glob), key=lambda p: p.stat().st_mtime, reverse=True)
    return matches[0] if matches else None


def find_inputs(folder: Path) -> dict:
    return {
        "county":   _latest(folder, "Absentee County*.pdf"),
        "cd":       _latest(folder, "Absentee Congressional*.pdf"),
        "voter_reg": _latest(folder, "voters by county*.pdf"),
    }

# ─────────────────────────────────────────────────────────────────────────────
# PDF PARSING
# ─────────────────────────────────────────────────────────────────────────────

def _normalize(text: str) -> str:
    return text.replace("’", "'").replace("‘", "'")


def parse_county_pdf(pdf_path: Path) -> pd.DataFrame:
    records = []
    current_county = None
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            raw = page.extract_text()
            if not raw:
                continue
            for line in _normalize(raw).splitlines():
                line = line.strip()
                if not line:
                    continue
                parts = line.split()
                nums, name_parts = [], []
                for p in parts:
                    try:
                        nums.append(int(p.replace(",", "")))
                    except ValueError:
                        name_parts.append(p)
                name = " ".join(name_parts)
                if name in IOWA_COUNTIES and len(nums) == 3:
                    current_county = name
                elif name in _ALL_PARTIES and len(nums) in (2, 3) and current_county:
                    party = name if name in _MAIN_PARTIES else "Other"
                    records.append({
                        "county": current_county,
                        "party": party,
                        "requested": nums[0],
                        "issued":    nums[1],
                        "received":  nums[2] if len(nums) == 3 else 0,
                    })
    return pd.DataFrame(records)


def parse_cd_pdf(pdf_path: Path) -> pd.DataFrame:
    records = []
    current_cd = None
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            raw = page.extract_text()
            if not raw:
                continue
            for line in _normalize(raw).splitlines():
                line = line.strip()
                if not line:
                    continue
                parts = line.split()
                nums, name_parts = [], []
                for p in parts:
                    try:
                        nums.append(int(p.replace(",", "")))
                    except ValueError:
                        name_parts.append(p)
                name = " ".join(name_parts)
                if name.startswith("United States Representative District") and len(nums) == 4:
                    current_cd = f"CD{nums[0]}"
                elif name in _ALL_PARTIES and len(nums) in (2, 3) and current_cd:
                    party = name if name in _MAIN_PARTIES else "Other"
                    records.append({
                        "cd": current_cd,
                        "party": party,
                        "requested": nums[0],
                        "issued":    nums[1],
                        "received":  nums[2] if len(nums) == 3 else 0,
                    })
    return pd.DataFrame(records)


def parse_voter_reg_pdf(pdf_path: Path) -> pd.DataFrame:
    """
    Returns DataFrame: county, dem_active, rep_active, np_active, total_active
    PDF columns (after county name): DemA RepA NPA OthA TotA DemI RepI NPI OthI TotI Grand
    """
    records = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            raw = page.extract_text()
            if not raw:
                continue
            for line in _normalize(raw).splitlines():
                line = line.strip()
                if not line:
                    continue
                parts = line.split()
                # County name may be multi-word; scan from left until we find a run of nums
                nums, name_parts = [], []
                for p in parts:
                    try:
                        nums.append(int(p.replace(",", "")))
                    except ValueError:
                        if not nums:  # still in the name portion
                            name_parts.append(p)
                name = " ".join(name_parts)
                if name in IOWA_COUNTIES and len(nums) >= 5:
                    records.append({
                        "county":      name,
                        "dem_active":  nums[0],
                        "rep_active":  nums[1],
                        "np_active":   nums[2],
                        "total_active": nums[4],
                    })
    return pd.DataFrame(records)

# ─────────────────────────────────────────────────────────────────────────────
# AGGREGATION — standard ballot summary
# ─────────────────────────────────────────────────────────────────────────────

BALLOT_COLS = [
    "Geography",
    "Dem Requested", "Dem Issued", "Dem Received",
    "Rep Requested", "Rep Issued", "Rep Received",
    "Other Requested", "Other Issued", "Other Received",
    "Total Requested", "Total Issued", "Total Received",
]


def _party_agg(df: pd.DataFrame, group_col: str, party: str, label: str) -> pd.DataFrame:
    return (
        df[df["party"] == party]
        .groupby(group_col)[["requested", "issued", "received"]]
        .sum()
        .rename(columns={
            "requested": f"{label} Requested",
            "issued":    f"{label} Issued",
            "received":  f"{label} Received",
        })
    )


def build_summary(df: pd.DataFrame, group_col: str) -> pd.DataFrame:
    dem   = _party_agg(df, group_col, "Democrat",   "Dem")
    rep   = _party_agg(df, group_col, "Republican", "Rep")
    other = _party_agg(df, group_col, "Other",      "Other")
    combined = dem.join(rep, how="outer").join(other, how="outer").fillna(0).astype(int)
    combined["Total Requested"] = combined["Dem Requested"] + combined["Rep Requested"] + combined["Other Requested"]
    combined["Total Issued"]    = combined["Dem Issued"]    + combined["Rep Issued"]    + combined["Other Issued"]
    combined["Total Received"]  = combined["Dem Received"]  + combined["Rep Received"]  + combined["Other Received"]
    combined = combined.reset_index().rename(columns={group_col: "Geography"}).sort_values("Geography").reset_index(drop=True)
    totals = {c: combined[c].sum() for c in combined.columns if c != "Geography"}
    totals["Geography"] = "TOTAL"
    combined = pd.concat([combined, pd.DataFrame([totals])], ignore_index=True)
    return combined[BALLOT_COLS]

# ─────────────────────────────────────────────────────────────────────────────
# SHARE ANALYSIS
# ─────────────────────────────────────────────────────────────────────────────

SHARE_COLS = [
    "Market",
    "GOP Received", "GOP Early %", "GOP Reg %", "GOP Over/Under",
    "Dem Received", "Dem Early %", "Dem Reg %", "Dem Over/Under",
    "Total Received",
]


def build_share_analysis(raw: pd.DataFrame, voter_reg: pd.DataFrame) -> pd.DataFrame:
    # County-level early votes received
    county_votes = (
        raw[raw["party"].isin(["Democrat", "Republican"])]
        .pivot_table(index="county", columns="party", values="received", aggfunc="sum", fill_value=0)
        .rename(columns={"Democrat": "dem_received", "Republican": "rep_received"})
        .reset_index()
    )
    county_votes["total_received"] = county_votes["dem_received"] + county_votes["rep_received"]

    # Add market
    county_votes["market"] = county_votes["county"].map(COUNTY_TO_MARKET)

    # Merge registration
    merged = county_votes.merge(voter_reg[["county", "dem_active", "rep_active"]], on="county", how="left")

    # Statewide totals
    tot_gop_rec = merged["rep_received"].sum()
    tot_dem_rec = merged["dem_received"].sum()
    tot_gop_reg = merged["rep_active"].sum()
    tot_dem_reg = merged["dem_active"].sum()

    # Aggregate by market
    mkt = merged.groupby("market").agg(
        gop_received=("rep_received", "sum"),
        dem_received=("dem_received", "sum"),
        total_received=("total_received", "sum"),
        gop_reg=("rep_active", "sum"),
        dem_reg=("dem_active", "sum"),
    ).reset_index()

    mkt["GOP Early %"]   = mkt["gop_received"] / tot_gop_rec * 100
    mkt["GOP Reg %"]     = mkt["gop_reg"]      / tot_gop_reg * 100
    mkt["GOP Over/Under"]= mkt["GOP Early %"]  - mkt["GOP Reg %"]
    mkt["Dem Early %"]   = mkt["dem_received"] / tot_dem_rec * 100
    mkt["Dem Reg %"]     = mkt["dem_reg"]      / tot_dem_reg * 100
    mkt["Dem Over/Under"]= mkt["Dem Early %"]  - mkt["Dem Reg %"]

    mkt = mkt.rename(columns={
        "market": "Market",
        "gop_received": "GOP Received",
        "dem_received": "Dem Received",
        "total_received": "Total Received",
    }).sort_values("Market").reset_index(drop=True)

    # Totals row
    totals_row = {
        "Market": "TOTAL",
        "GOP Received": int(mkt["GOP Received"].sum()),
        "GOP Early %":   100.0,
        "GOP Reg %":     100.0,
        "GOP Over/Under": 0.0,
        "Dem Received": int(mkt["Dem Received"].sum()),
        "Dem Early %":   100.0,
        "Dem Reg %":     100.0,
        "Dem Over/Under": 0.0,
        "Total Received": int(mkt["Total Received"].sum()),
    }
    mkt = pd.concat([mkt, pd.DataFrame([totals_row])], ignore_index=True)
    return mkt[SHARE_COLS]

# ─────────────────────────────────────────────────────────────────────────────
# COUNTY-LEVEL OVER/UNDER (for map)
# ─────────────────────────────────────────────────────────────────────────────

def build_county_overunder(raw: pd.DataFrame, voter_reg: pd.DataFrame) -> pd.DataFrame:
    county_votes = (
        raw[raw["party"].isin(["Democrat", "Republican"])]
        .pivot_table(index="county", columns="party", values="received", aggfunc="sum", fill_value=0)
        .rename(columns={"Democrat": "dem_received", "Republican": "rep_received"})
        .reset_index()
    )
    merged = county_votes.merge(voter_reg[["county", "dem_active", "rep_active"]], on="county", how="left")

    tot_gop_rec = merged["rep_received"].sum()
    tot_dem_rec = merged["dem_received"].sum()
    tot_gop_reg = merged["rep_active"].sum()
    tot_dem_reg = merged["dem_active"].sum()

    merged["gop_early_pct"] = merged["rep_received"] / tot_gop_rec * 100
    merged["gop_reg_pct"]   = merged["rep_active"]   / tot_gop_reg * 100
    merged["gop_overunder"] = merged["gop_early_pct"] - merged["gop_reg_pct"]

    merged["dem_early_pct"] = merged["dem_received"] / tot_dem_rec * 100
    merged["dem_reg_pct"]   = merged["dem_active"]   / tot_dem_reg * 100
    merged["dem_overunder"] = merged["dem_early_pct"] - merged["dem_reg_pct"]

    merged["fips"]   = merged["county"].map(COUNTY_TO_FIPS)
    merged["market"] = merged["county"].map(COUNTY_TO_MARKET)
    return merged

# ─────────────────────────────────────────────────────────────────────────────
# HTML MAP
# ─────────────────────────────────────────────────────────────────────────────

GEOJSON_URL  = "https://raw.githubusercontent.com/plotly/datasets/master/geojson-counties-fips.json"
GEOJSON_CACHE = Path(__file__).parent / "counties_fips.json"


def _load_geojson() -> dict:
    if GEOJSON_CACHE.exists():
        return json.loads(GEOJSON_CACHE.read_text())
    print("  Downloading county GeoJSON (one-time)…")
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    with urlopen(GEOJSON_URL, context=ctx) as r:
        data = json.load(r)
    GEOJSON_CACHE.write_text(json.dumps(data))
    return data


def _build_market_boundaries(geojson: dict) -> dict:
    """Return {market_name: shapely_geom} — dissolved county polygons per DMA."""
    from shapely.geometry import shape
    from shapely.ops import unary_union

    fips_to_market = {fips: COUNTY_TO_MARKET[county]
                      for county, fips in COUNTY_TO_FIPS.items()
                      if county in COUNTY_TO_MARKET}
    market_shapes: dict[str, list] = {}
    for feature in geojson["features"]:
        market = fips_to_market.get(feature["id"])
        if market:
            market_shapes.setdefault(market, []).append(shape(feature["geometry"]))
    return {market: unary_union(shapes) for market, shapes in market_shapes.items()}


def _make_choropleth(df: pd.DataFrame, geojson: dict, value_col: str,
                     hover_cols: list[str], title: str,
                     market_boundaries: dict | None = None,
                     hover_labels: list[str] | None = None) -> go.Figure:
    iowa_fips = set(COUNTY_TO_FIPS.values())
    features = [f for f in geojson["features"] if f["id"] in iowa_fips]
    iowa_geo = {**geojson, "features": features}

    abs_max = df[value_col].abs().max()
    if abs_max == 0:
        abs_max = 1.0

    labels = hover_labels if hover_labels is not None else hover_cols
    customdata = df[["market"] + hover_cols].values
    hover_template = (
        "<b>%{text}</b><br>"
        "Market: %{customdata[0]}<br>" +
        "".join(f"{lbl}: %{{customdata[{i+1}]}}<br>" for i, lbl in enumerate(labels)) +
        "<extra></extra>"
    )

    fig = go.Figure(go.Choropleth(
        geojson=iowa_geo,
        locations=df["fips"],
        z=df[value_col],
        text=df["county"],
        customdata=customdata,
        hovertemplate=hover_template,
        colorscale=[[0, "#d73027"], [0.5, "#f7f7f7"], [1, "#1a9850"]],
        zmid=0,
        zmin=-abs_max,
        zmax=abs_max,
        colorbar=dict(
            title=dict(text="Over/Under<br>(pct pts)", side="right"),
            thickness=15,
        ),
        marker_line_color="white",
        marker_line_width=0.5,
    ))

    # Overlay media market boundary lines
    if market_boundaries:
        from shapely.geometry import MultiPolygon, Polygon
        for geom in market_boundaries.values():
            polys = geom.geoms if isinstance(geom, MultiPolygon) else [geom]
            for poly in polys:
                coords = list(poly.exterior.coords)
                lons = [c[0] for c in coords]
                lats = [c[1] for c in coords]
                fig.add_trace(go.Scattergeo(
                    lon=lons, lat=lats,
                    mode="lines",
                    line=dict(color="#1f1f1f", width=2.5),
                    showlegend=False,
                    hoverinfo="skip",
                ))

    fig.update_geos(
        scope="usa",
        fitbounds="locations",
        visible=False,
    )
    fig.update_layout(
        title=dict(text=title, x=0.5, xanchor="center"),
        margin=dict(l=0, r=0, t=50, b=0),
        height=550,
    )
    return fig


def build_html_map(raw: pd.DataFrame, voter_reg: pd.DataFrame, out_path: Path,
                   share_df: pd.DataFrame | None = None) -> None:
    import json as _json
    import re as _re

    if share_df is None:
        share_df = build_share_analysis(raw, voter_reg)

    df = build_county_overunder(raw, voter_reg)
    geojson = _load_geojson()
    market_boundaries = _build_market_boundaries(geojson)

    def ou_str(v):
        return f"{'+' if v >= 0 else ''}{v:.2f} pp"

    # ── County-level label columns ────────────────────────────────────────────
    df["gop_ou_label"]   = df["gop_overunder"].apply(ou_str)
    df["dem_ou_label"]   = df["dem_overunder"].apply(ou_str)
    df["gop_early_fmt"]  = df["gop_early_pct"].apply(lambda v: f"{v:.2f}%")
    df["gop_reg_fmt"]    = df["gop_reg_pct"].apply(lambda v: f"{v:.2f}%")
    df["dem_early_fmt"]  = df["dem_early_pct"].apply(lambda v: f"{v:.2f}%")
    df["dem_reg_fmt"]    = df["dem_reg_pct"].apply(lambda v: f"{v:.2f}%")

    # ── Market-level values mapped back to each county ────────────────────────
    mkt = share_df[share_df["Market"] != "TOTAL"].set_index("Market")
    mdf = df[["county", "fips", "market"]].copy()
    for col, src in [
        ("gop_mkt_ou",    "GOP Over/Under"), ("dem_mkt_ou",    "Dem Over/Under"),
        ("gop_mkt_early", "GOP Early %"),    ("dem_mkt_early", "Dem Early %"),
        ("gop_mkt_reg",   "GOP Reg %"),      ("dem_mkt_reg",   "Dem Reg %"),
    ]:
        mdf[col] = mdf["market"].map(mkt[src])
    mdf["gop_mkt_ou_label"]    = mdf["gop_mkt_ou"].apply(ou_str)
    mdf["dem_mkt_ou_label"]    = mdf["dem_mkt_ou"].apply(ou_str)
    mdf["gop_mkt_early_fmt"]   = mdf["gop_mkt_early"].apply(lambda v: f"{v:.2f}%")
    mdf["gop_mkt_reg_fmt"]     = mdf["gop_mkt_reg"].apply(lambda v: f"{v:.2f}%")
    mdf["dem_mkt_early_fmt"]   = mdf["dem_mkt_early"].apply(lambda v: f"{v:.2f}%")
    mdf["dem_mkt_reg_fmt"]     = mdf["dem_mkt_reg"].apply(lambda v: f"{v:.2f}%")

    # ── Build 4 figures ───────────────────────────────────────────────────────
    _county_labels = ["Early %", "Reg %", "Over/Under"]
    _market_labels = ["Mkt Early %", "Mkt Reg %", "Over/Under"]
    figs = {
        "gop-county": _make_choropleth(df,  geojson, "gop_overunder",
            ["gop_early_fmt", "gop_reg_fmt", "gop_ou_label"],
            "GOP Early Vote Over/Under — by County", market_boundaries, _county_labels),
        "gop-market": _make_choropleth(mdf, geojson, "gop_mkt_ou",
            ["gop_mkt_early_fmt", "gop_mkt_reg_fmt", "gop_mkt_ou_label"],
            "GOP Early Vote Over/Under — by Media Market", market_boundaries, _market_labels),
        "dem-county": _make_choropleth(df,  geojson, "dem_overunder",
            ["dem_early_fmt", "dem_reg_fmt", "dem_ou_label"],
            "Dem Early Vote Over/Under — by County", market_boundaries, _county_labels),
        "dem-market": _make_choropleth(mdf, geojson, "dem_mkt_ou",
            ["dem_mkt_early_fmt", "dem_mkt_reg_fmt", "dem_mkt_ou_label"],
            "Dem Early Vote Over/Under — by Media Market", market_boundaries, _market_labels),
    }

    # ── Serialize to HTML fragments ───────────────────────────────────────────
    first_key = "gop-county"
    first_full = figs[first_key].to_html(full_html=False, include_plotlyjs="cdn",
                                          div_id=f"map-{first_key}")
    cdn_match = _re.search(r'<script[^>]*src="[^"]*plotly[^"]*"[^>]*></script>', first_full)
    cdn_tag = cdn_match.group(0) if cdn_match else '<script src="https://cdn.plot.ly/plotly-latest.min.js"></script>'

    panels_html = ""
    for key, fig in figs.items():
        if key == first_key:
            raw_html = first_full.replace(cdn_tag, "")
        else:
            raw_html = fig.to_html(full_html=False, include_plotlyjs=False, div_id=f"map-{key}")
        active = " active" if key == first_key else ""
        panels_html += f'<div id="panel-{key}" class="map-panel{active}">{raw_html}</div>\n'

    # ── Sidebar table data ────────────────────────────────────────────────────
    rows = share_df[share_df["Market"] != "TOTAL"]
    gop_table = [{"market": r["Market"], "early": f"{r['GOP Early %']:.2f}%",
                  "reg": f"{r['GOP Reg %']:.2f}%", "ou": ou_str(r["GOP Over/Under"]),
                  "pos": r["GOP Over/Under"] >= 0} for _, r in rows.iterrows()]
    dem_table = [{"market": r["Market"], "early": f"{r['Dem Early %']:.2f}%",
                  "reg": f"{r['Dem Reg %']:.2f}%", "ou": ou_str(r["Dem Over/Under"]),
                  "pos": r["Dem Over/Under"] >= 0} for _, r in rows.iterrows()]
    table_json = _json.dumps({"gop": gop_table, "dem": dem_table})

    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>Iowa Early Vote Over/Under Map</title>
{cdn_tag}
<style>
  * {{ box-sizing: border-box; }}
  body {{ font-family: Arial, sans-serif; margin: 0; background: #f5f5f5; }}
  h1 {{ text-align: center; color: #1f3a6e; padding: 16px 0 4px; margin: 0; font-size: 1.3em; }}
  .toggle-row {{ display: flex; justify-content: center; gap: 12px; padding: 4px 0; }}
  .toggle-row.party {{ padding-top: 8px; }}
  .toggle-row.view  {{ padding-bottom: 10px; }}
  .tab-btn, .view-btn {{
    padding: 7px 24px; font-size: 0.95em; border: 2px solid #1f3a6e;
    border-radius: 4px; cursor: pointer; background: white; color: #1f3a6e; font-weight: bold;
  }}
  .tab-btn.active, .view-btn.active {{ background: #1f3a6e; color: white; }}
  .note {{ text-align: center; color: #666; font-size: 0.85em; padding-bottom: 8px; }}
  .content {{ display: flex; align-items: flex-start; padding: 0 12px 16px; gap: 14px; }}
  .map-wrap {{ flex: 1; min-width: 0; }}
  .map-panel {{ visibility: hidden; height: 0; overflow: hidden; }}
  .map-panel.active {{ visibility: visible; height: auto; overflow: visible; }}
  .sidebar {{ width: 310px; flex-shrink: 0; }}
  .sidebar h2 {{ font-size: 0.9em; color: #1f3a6e; margin: 0 0 6px; text-align: center; }}
  .mkt-table {{ width: 100%; border-collapse: collapse; font-size: 0.82em; }}
  .mkt-table th {{ background: #1f3a6e; color: white; padding: 6px 8px; text-align: left; }}
  .mkt-table td {{ padding: 5px 8px; border-bottom: 1px solid #ddd; }}
  .mkt-table tr:last-child td {{ border-bottom: none; }}
  .mkt-table tr:nth-child(even) td {{ background: #eef2f8; }}
  .ou-pos {{ color: #1a9850; font-weight: bold; }}
  .ou-neg {{ color: #d73027; font-weight: bold; }}
</style>
</head>
<body>
<h1>Iowa Early Vote Over/Under vs. Registered Voter Share</h1>
<div class="toggle-row party">
  <button class="tab-btn active"  onclick="setParty('gop', this)">GOP</button>
  <button class="tab-btn"         onclick="setParty('dem', this)">Dem</button>
</div>
<div class="toggle-row view">
  <button class="view-btn active" onclick="setView('county', this)">County</button>
  <button class="view-btn"        onclick="setView('market', this)">Market</button>
</div>
<p class="note">Green = over-performing early vote share &nbsp;|&nbsp; Red = under-performing &nbsp;|&nbsp; Hover for details</p>
<div class="content">
  <div class="map-wrap">
{panels_html}  </div>
  <div class="sidebar">
    <h2 id="tbl-title">GOP — Media Market Share</h2>
    <table class="mkt-table">
      <thead><tr><th>Market</th><th>Early %</th><th>Reg %</th><th>Over/Under</th></tr></thead>
      <tbody id="tbl-body"></tbody>
    </table>
  </div>
</div>
<script>
var TABLE_DATA = {table_json};
var currentParty = 'gop';
var currentView  = 'county';
function renderTable(party) {{
  var rows = TABLE_DATA[party];
  var titles = {{gop: 'GOP', dem: 'Dem'}};
  document.getElementById('tbl-title').textContent = titles[party] + ' — Media Market Share';
  var h = '';
  rows.forEach(function(r) {{
    var cls = r.pos ? 'ou-pos' : 'ou-neg';
    h += '<tr><td>' + r.market + '</td><td>' + r.early + '</td><td>' + r.reg +
         '</td><td class="' + cls + '">' + r.ou + '</td></tr>';
  }});
  document.getElementById('tbl-body').innerHTML = h;
}}
function _showPanel() {{
  document.querySelectorAll('.map-panel').forEach(function(p) {{ p.classList.remove('active'); }});
  var id = 'panel-' + currentParty + '-' + currentView;
  document.getElementById(id).classList.add('active');
  Plotly.relayout('map-' + currentParty + '-' + currentView, {{autosize: true}});
  renderTable(currentParty);
}}
function setParty(party, btn) {{
  currentParty = party;
  document.querySelectorAll('.tab-btn').forEach(function(b) {{ b.classList.remove('active'); }});
  btn.classList.add('active');
  _showPanel();
}}
function setView(view, btn) {{
  currentView = view;
  document.querySelectorAll('.view-btn').forEach(function(b) {{ b.classList.remove('active'); }});
  btn.classList.add('active');
  _showPanel();
}}
renderTable('gop');
</script>
</body>
</html>"""

    out_path.write_text(html, encoding="utf-8")
    print(f"  Map saved: {out_path}")

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL OUTPUT
# ─────────────────────────────────────────────────────────────────────────────

HEADER_BLUE = "1F4E79"
TOTALS_GRAY = "D9D9D9"
GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def _style_ballot_sheet(ws, total_rows: int) -> None:
    header_fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type="solid")
    totals_fill = PatternFill(start_color=TOTALS_GRAY, end_color=TOTALS_GRAY, fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for cell in ws[total_rows + 2]:
        cell.font = Font(bold=True)
        cell.fill = totals_fill
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.column > 1:
                cell.number_format = "#,##0"
    for col in ws.columns:
        w = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(w + 3, 30)


def _style_share_sheet(ws, df: pd.DataFrame) -> None:
    header_fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type="solid")
    totals_fill = PatternFill(start_color=TOTALS_GRAY, end_color=TOTALS_GRAY, fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    totals_row_idx = len(df) + 1  # 1-indexed, +1 for header
    for cell in ws[totals_row_idx + 1]:
        cell.font = Font(bold=True)
        cell.fill = totals_fill

    # Column index lookup
    headers = [cell.value for cell in ws[1]]
    pct_cols   = {h: i+1 for i, h in enumerate(headers) if "%" in str(h)}
    ou_cols    = {h: i+1 for i, h in enumerate(headers) if "Over/Under" in str(h)}
    count_cols = {h: i+1 for i, h in enumerate(headers) if h in ("GOP Received","Dem Received","Total Received")}

    for row in ws.iter_rows(min_row=2, max_row=len(df) + 1):
        for cell in row:
            h = headers[cell.column - 1]
            if h in pct_cols:
                cell.number_format = "0.00%"
                if isinstance(cell.value, (int, float)):
                    cell.value = cell.value / 100
            elif h in ou_cols:
                cell.number_format = '+0.00%;-0.00%;"0.00%"'
                if isinstance(cell.value, (int, float)):
                    v = cell.value / 100
                    cell.value = v
                    cell.fill = GREEN_FILL if v > 0 else (RED_FILL if v < 0 else PatternFill())
            elif h in count_cols:
                cell.number_format = "#,##0"

    # Totals row: same pct/ou formatting but no conditional color
    for cell in ws[totals_row_idx + 1]:
        h = headers[cell.column - 1] if cell.column <= len(headers) else ""
        if h in pct_cols or h in ou_cols:
            if isinstance(cell.value, (int, float)):
                cell.value = cell.value / 100
            cell.number_format = "0.00%"
        elif h in count_cols:
            cell.number_format = "#,##0"

    for col in ws.columns:
        w = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, 28)


def write_excel(market_df, cd_df, county_df, share_df, out_path: Path) -> None:
    sheets = [
        (market_df, "By Media Market",          "ballot"),
        (cd_df,     "By Congressional District","ballot"),
        (county_df, "By County",                "ballot"),
        (share_df,  "Market Share Analysis",    "share"),
    ]
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for df, name, kind in sheets:
            df.to_excel(writer, sheet_name=name, index=False)
            ws = writer.sheets[name]
            if kind == "ballot":
                _style_ballot_sheet(ws, len(df) - 1)
            else:
                _style_share_sheet(ws, df)
    print(f"  Excel saved: {out_path}")

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main(county_pdf_override: Path | None = None) -> None:
    base = Path(__file__).parent
    inputs = find_inputs(base / "data")

    county_pdf = county_pdf_override or inputs["county"]
    cd_pdf     = inputs["cd"]
    reg_pdf    = inputs["voter_reg"]

    if not county_pdf or not county_pdf.exists():
        raise FileNotFoundError("No 'Absentee County*.pdf' found in folder.")
    if not reg_pdf or not reg_pdf.exists():
        raise FileNotFoundError("No 'voters by county*.pdf' found in folder.")

    stem     = county_pdf.stem.replace(" ", "_")
    out_dir  = base / "output"
    out_dir.mkdir(exist_ok=True)
    out_xlsx = out_dir / f"{stem}_Breakdown.xlsx"
    out_map  = out_dir / f"{stem}_map.html"

    print(f"\n{'='*60}")
    print(f"County PDF : {county_pdf.name}")
    print(f"CD PDF     : {cd_pdf.name if cd_pdf else 'not found'}")
    print(f"Voter Reg  : {reg_pdf.name}")
    print(f"Output     : {out_xlsx.name}  +  {out_map.name}")
    print(f"{'='*60}")

    # Parse
    print("Parsing county PDF…")
    raw = parse_county_pdf(county_pdf)
    print(f"  {raw['county'].nunique()} counties, {len(raw)} party records")
    print(f"  Grand Total Requested: {raw['requested'].sum():,}")
    print(f"  Grand Total Received:  {raw['received'].sum():,}")

    print("Parsing voter registration PDF…")
    voter_reg = parse_voter_reg_pdf(reg_pdf)
    print(f"  {len(voter_reg)} counties parsed")
    print(f"  Total active GOP reg: {voter_reg['rep_active'].sum():,}")
    print(f"  Total active Dem reg: {voter_reg['dem_active'].sum():,}")

    # Apply market mapping
    raw["market"] = raw["county"].map(COUNTY_TO_MARKET)

    # Build ballot summary sheets
    print("Building ballot summaries…")
    market_df = build_summary(raw, "market")

    if cd_pdf and cd_pdf.exists():
        print("Parsing congressional district PDF…")
        cd_raw = parse_cd_pdf(cd_pdf)
        cd_df  = build_summary(cd_raw, "cd")
        print(f"  CD grand total: {cd_df.loc[cd_df['Geography']=='TOTAL','Total Requested'].values[0]:,}")
    else:
        print("  WARNING — CD PDF not found; using county mapping (approximate).")
        raw["cd"] = raw["county"].map(COUNTY_TO_CD)
        cd_df = build_summary(raw, "cd")

    county_df = build_summary(raw, "county")
    county_df.insert(1, "Media Market",   county_df["Geography"].map(COUNTY_TO_MARKET).fillna(""))
    county_df.insert(2, "Cong. District", county_df["Geography"].map(COUNTY_TO_CD).fillna(""))

    # Build share analysis sheet
    print("Building market share analysis…")
    share_df = build_share_analysis(raw, voter_reg)

    # Write Excel
    print("Writing Excel workbook…")
    write_excel(market_df, cd_df, county_df, share_df, out_xlsx)

    # Generate map
    print("Generating HTML map…")
    build_html_map(raw, voter_reg, out_map, share_df=share_df)

    print("Done.")


if __name__ == "__main__":
    main()
